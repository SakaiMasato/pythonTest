#! python3
# -*- coding: utf-8 -*-

' excel stuff '

__author__ = 'Bob Bao'

import openpyxl, logging
from openpyxl.utils import get_column_letter, column_index_from_string
from openpyxl.styles import Font, PatternFill, colors
logging.basicConfig(level=logging.DEBUG, format=' %(asctime)s - %(levelname)s- %(message)s')

wb = openpyxl.load_workbook('bob.xlsx')
logging.info(type(wb))
sheet = wb.worksheets[0]                                                                            #first sheet
logging.info('first worksheet: ' + sheet.title)
logging.info('first cell: value-' + sheet['A1'].value + ' row-' + str(sheet['A1'].row) + ' column-' + str(sheet['A1'].column) + ' coordinate-' + sheet['A1'].coordinate)
logging.info('first cell: value-' + sheet.cell(1,1).value+ ' row-' + str(sheet.cell(1,1).row) + ' column-' + str(sheet.cell(1,1).column) + ' coordinate-' + sheet.cell(1,1).coordinate)
logging.info('len(row): ' + str(sheet.max_row))
logging.info('len(column): ' + str(sheet.max_column))
logging.info('translate int to str: ' + get_column_letter(sheet.max_row))
logging.info('translate str to int: ' + str(column_index_from_string('B')))
logging.info(tuple(sheet['A1': 'C3']))

logging.info('=================excel write==================')
wb1 = openpyxl.Workbook()       ##create a new workbook
sheet1 = wb1.worksheets[0]
logging.info('sheet title: ' + sheet1.title)        #default sheet name is 'Sheet'
sheet1.title = 'Spam Bacon Eggs sheet'              ##Modify first sheet name
logging.info('sheet title: ' + sheet1.title)
wb1.save('bobTestExcel.xlsx')

wb2 = openpyxl.Workbook()
wb2.create_sheet()
wb2.create_sheet(index=0, title='1st sheet')
wb2.create_sheet(index=2, title='middle sheet')
sheet2 = wb2['middle sheet']
sheet2['A1'].font = Font(size=24, italic=True)      #font
sheet2['A1'].fill = PatternFill(fill_type='solid', fgColor=colors.GREEN)      #fill
sheet2.merge_cells('A1:D3')                         #merge cells
sheet2.unmerge_cells('A1:D3')                       #unmerge_cells
'''
sheet.freeze_panes = 'A2' 行1
sheet.freeze_panes = 'B1' 列A
sheet.freeze_panes = 'C1' 列A 和列B
sheet.freeze_panes = 'C2' 行1 和列A 和列B
sheet.freeze_panes = 'A1'或
sheet.freeze_panes = None
'''
sheet2.freeze_panes = 'A2'                          #freeze row 1
sheet2.cell(1,1).value = 'Hello world'
sheet2.row_dimensions[1].height = 70                #first row height
sheet2.column_dimensions['A'].width = 20            #first col width
logging.info(sheet2['A1'].value)

#formula sheet['B9'] = '=SUM(B1:B8)'


#############################graph###############################################
sheet3 = wb2['1st sheet']
for i in range(1, 11):
    sheet3['A' + str(i)] = i
refObj = openpyxl.chart.Reference(sheet3, min_row=1, min_col=1, max_row=10, max_col=1)
seriesObj = openpyxl.chart.Series(refObj, title='First series')
chartObj = openpyxl.chart.BarChart()
chartObj.title = 'My Chart'
chartObj.append(seriesObj)
sheet3.add_chart(chartObj, 'C5')
wb2.save('bobTestExcel1.xlsx')