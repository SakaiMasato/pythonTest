#! python3
# -*- coding: utf-8 -*-

' update excel info '

__author__ = 'Bob Bao'

import openpyxl, logging
logging.basicConfig(level=logging.DEBUG, format=' %(asctime)s - %(levelname)s- %(message)s')

updateItem = {
    'Celery' : 1.19,
    'Garlic' : 3.07,
    'Lemon' : 1.27
}

nameCol = 'FRUIT'
singlePriceCol = 'COST PER POUND'
nameColNum = None
singlePriceColNum = None

if __name__ == '__main__':
    wb = openpyxl.load_workbook('freezeExample.xlsx')
    sheet = wb.worksheets[0]
    max_row = sheet.max_row
    max_col = sheet.max_column
    for i in range(1, max_col):
        val = sheet.cell(1, i).value
        if(val == nameCol):
            nameColNum = i
        elif(val == singlePriceCol):
            singlePriceColNum = i

    for i in range(1, max_row):
        if sheet.cell(i, nameColNum).value in updateItem:
            sheet.cell(i, singlePriceColNum).value = updateItem[sheet.cell(i, nameColNum).value]

    wb.save('updatedFreezeExample.xlsx')