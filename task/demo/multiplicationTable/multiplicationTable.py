import openpyxl, logging
from  openpyxl.styles import  PatternFill, colors, Font
logging.basicConfig(level=logging.DEBUG, format=' %(asctime)s - %(levelname)s- %(message)s')

wb = openpyxl.Workbook()
sheet = wb.active
sheet.title = 'muitiplicationTable'
for i in range(1, 10):
    sheet['A'+str(i+1)] = i
    sheet['A'+str(i+1)].fill = PatternFill(fill_type='solid', fgColor=colors.RED)
    sheet['A' + str(i + 1)].font = Font(bold=True)
for i in range(1, 10):
    sheet.cell(1, i+1).value = i
    sheet.cell(1, i + 1).fill = PatternFill(fill_type='solid', fgColor=colors.RED)
    sheet.cell(1, i + 1).font = Font(bold=True)
for i in range(1, 10):
    for j in range(1,10):
        sheet.cell(i+1,j+1).value = sheet.cell(i+1,1).value * sheet.cell(1,j+1).value
wb.save('multiplicationTable.xlsx')