#! python3
# -*- coding: utf-8 -*-

' read from excel '

__author__ = 'Bob Bao'

import openpyxl,logging
logging.basicConfig(level=logging.DEBUG, format=' %(asctime)s - %(levelname)s- %(message)s')

counties = {}
sumPeople = 0
excelName = 'censuspopdata.xlsx'

def calNumber(workbook, filter, numCol):
    rowLen = workbook.max_row
    colLen = workbook.max_column
    filterColumn = ''
    group = {}
    for i in range(1, colLen):
        if(sheet.cell(1, i).value == filter):
            filterColumn = i                              #get which column to group
    if (filterColumn != ''):
        for i in range(2, rowLen):                        #start from 2 hence 1 line is the title
            group.setdefault(sheet.cell(i, filterColumn).value, 0)
            if(numCol is not None):
                group[sheet.cell(i, filterColumn).value] += int(sheet.cell(i, sumColumn).value)
            else:
                group[sheet.cell(i, filterColumn).value] += 1    #default operation is to sum up the length
    return group;

if __name__ == '__main__':
    wb = openpyxl.load_workbook(excelName)
    sheet = wb.get_active_sheet()
    rowLen = sheet.max_row
    colLen = sheet.max_column
    sumColumn = 4                                       #summary column is 4
    #print(rowLen)
    #print(colLen)
    countyGroup = calNumber(sheet, 'County', sumColumn)
    CensusTractGroup = calNumber(sheet, 'County', None)
    logging.info('groupby County: ' + str(countyGroup['Sheridan']))
    logging.info('CensusTract number in CA state: ' + str(CensusTractGroup['Sheridan']))

